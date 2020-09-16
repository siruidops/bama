#!/usr/bin/env python3

import requests
from bs4 import BeautifulSoup
import html
import time
import datetime as dt
import sys
import os
import threading
from openpyxl import load_workbook
from openpyxl import Workbook

url_list = ['https://bama.ir/car','https://bama.ir/motorcycle']
urls = []

lock1 = threading.Lock()
lock2 = threading.Lock()
lock3 = threading.Lock()

datenow = dt.datetime.now()
year = datenow.year; month = datenow.month; day = datenow.day
timenow = "{}_{}_{}".format(year,month,day)

if not os.path.isfile('bama-car-{}.xlsx'.format(timenow)):
    workbook_car = Workbook()
    sh_car = workbook_car.active
    sh_car['A1'] = 'Group'
    sh_car['B1'] = 'Title'; sh_car.column_dimensions['B'].width = 15
    sh_car['C1'] = 'Url'
    sh_car['D1'] = 'Location'
    sh_car['E1'] = 'Time'
    sh_car['F1'] = 'Kilometre'
    sh_car['G1'] = 'Gearbox'
    sh_car['H1'] = 'Fuel'
    sh_car['I1'] = 'Body'
    sh_car['J1'] = 'Color'
    sh_car['K1'] = 'Description'; sh_car.column_dimensions['M'].width = 65
    sh_car['L1'] = 'Price'; 
    sh_car['M1'] = 'Pictures'; sh_car.column_dimensions['N'].width = 15

else:
    workbook_car = load_workbook('bama-car-{}.xlsx'.format(timenow))
    sh_car = workbook_car.worksheets[0]
    for i in range(2, sh_car.max_row+1):
        urls.append(sh_car.cell(row=i, column=3).value.strip())


if not os.path.isfile('bama-motor-{}.xlsx'.format(timenow)):
    workbook_motor = Workbook()
    sh_motor = workbook_motor.active
    sh_motor['A1'] = 'Group'
    sh_motor['B1'] = 'Title'; sh_motor.column_dimensions['A'].width = 15
    sh_motor['C1'] = 'Url'
    sh_motor['D1'] = 'Location'
    sh_motor['E1'] = 'Time'
    sh_motor['F1'] = 'Amount of worked'
    sh_motor['G1'] = 'Engine capacity'
    sh_motor['H1'] = 'Gearbox'
    sh_motor['I1'] = 'Fuel'
    sh_motor['J1'] = 'Color'
    sh_motor['K1'] = 'Description'; sh_motor.column_dimensions['M'].width = 65
    sh_motor['L1'] = 'Price'; 
    sh_motor['M1'] = 'Pictures'; sh_motor.column_dimensions['N'].width = 15

else:
    workbook_motor = load_workbook('bama-motor-{}.xlsx'.format(timenow))
    sh_motor = workbook_motor.worksheets[0]
    for i in range(2, sh_motor.max_row+1):
        urls.append(sh_motor.cell(row=i, column=3).value.strip())


def runner():
    url = url_list.pop()

    while 1:
        url_status_re = 0

        for i in range(1,41):
            if url_status_re == 1:
                break

            r = requests.Session()
            r.headers = {'Connection': "close", "Accept": "*/*", "Content-type": "application/x-www-form-urlencoded; charset=UTF-8", "Accept-Language": "en-US", "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"}
            rr = r.get("{}?page={}".format(url,str(i)))
            source_html = rr.text
            bs = BeautifulSoup(source_html,'html.parser')
            dom = bs.find_all("a", {'itemprop':'url'})

            for i in dom:
                l = i['href']
                if l not in urls:
                    urls.append(l)
                    url_status_re = 0

                else:
                    url_status_re = 1
                    continue


                k = r.get(l)
                bs_ = BeautifulSoup(k.text,'html.parser')
                title = bs_.find('h1', {'class':'addetail-title'}).text
                inforigth = bs_.find('div', {'class':'inforight'})
                price = inforigth.find('span', {'style':"color: #555;"})

                if price == None:
                    price = inforigth.find('span', {'content':'0'})
                    if price == None:
                        price = "اقساطی"

                if type(price) != str:
                    price = price.text

                time_ = ' '
                worked = ' '
                engine_size = ' '
                girbox = ' '
                sokht = ' '
                badane = ' '
                color = ' '
                ostan = ' '
                shahrestan = ' '
                mahale = ' '
                group = ' '
                bazdid = ' '

                lll = inforigth.find_all('span')
                image_url = []
                images = bs_.find_all('a', {'class':"bamalightgallery-item"})

                for image in images:
                    image_url.append(image['href'])

                pictures = ' , '.join(image_url)
                try:
                    description = bs_.find('span', {'class':"removeEmoji"}).text.strip()
                except:
                    description = ' '

                if url == "https://bama.ir/motorcycle":
                    group = 'موتور سیکلت'

                    for i in lll:
                        if i.text.strip() == 'زمان':
                            number = lll.index(i)
                            time_ = lll[number+1].text.strip()
                        elif i.text.strip() == 'كاركرد':
                            number = lll.index(i)
                            worked = lll[number+1].text.strip()
                        elif i.text.strip() == 'حجم موتور':
                            number = lll.index(i)
                            engine_size = "{} {}".format(lll[number+1].text.strip(), lll[number+2].text.strip())
                        elif i.text.strip() == 'گیربکس':
                            number = lll.index(i)
                            girbox = html.unescape(lll[number+1].text.strip()).strip()
                        elif i.text.strip() == 'سوخت':
                            number = lll.index(i)
                            sokht = html.unescape(lll[number+1].text.strip()).strip()
                        elif i.text.strip() == 'رنگ':
                            number = lll.index(i)
                            color = html.unescape(lll[number+1].text.strip()).strip()
                        elif i.text.strip() == 'استان':
                            number = lll.index(i)
                            ostan = html.unescape(lll[number+1].text.strip()).strip()
                        elif i.text.strip() == 'شهرستان':
                            number = lll.index(i)
                            shahrestan = html.unescape(lll[number+1].text.strip()).strip()
                        elif i.text.strip() == 'بازديد':
                            number = lll.index(i)
                            bazdid = html.unescape(lll[number+1].text.strip()).strip()
                        else:
                            pass

                    location = "{} - {} -{}".format(ostan, shahrestan, bazdid)
                    expens = [group, title, l, location, time_, worked, engine_size, girbox, sokht, color, description, price, pictures]
                    sh_motor.append(expens)
                    workbook_motor.save('bama-motor-{}.xlsx'.format(timenow))

                else:
                    group = 'خودرو'

                    for i in lll:
                        if i.text.strip() == 'زمان':
                            number = lll.index(i)
                            time_ = lll[number+1].text.strip()
                        elif i.text.strip() == 'كاركرد':
                            number = lll.index(i)
                            worked = lll[number+1].text.strip()
                        elif i.text.strip() == 'گیربکس':
                            number = lll.index(i)
                            girbox = html.unescape(lll[number+1].text.strip()).strip()
                        elif i.text.strip() == 'سوخت':
                            number = lll.index(i)
                            sokht = html.unescape(lll[number+1].text.strip()).strip()
                        elif i.text.strip() == 'بدنه':
                            number = lll.index(i)
                            badane = html.unescape(lll[number+1].text.strip()).strip()
                        elif i.text.strip() == 'رنگ':
                            number = lll.index(i)
                            color = html.unescape(lll[number+1].text.strip()).strip()
                        elif i.text.strip() == 'استان':
                            number = lll.index(i)
                            ostan = html.unescape(lll[number+1].text.strip()).strip()
                        elif i.text.strip() == 'شهرستان':
                            number = lll.index(i)
                            shahrestan = html.unescape(lll[number+1].text.strip()).strip()
                        elif i.text.strip() == 'محله':
                            number = lll.index(i)
                            mahale = html.unescape(lll[number+1].text.strip()).strip()
                        else:
                            pass

                    location = "{} - {} - {}".format(ostan, shahrestan, mahale)
                    expens = [group, title, l, location, time_, worked, girbox, sokht, badane, color, description, price, pictures]
                    sh_car.append(expens)
                    workbook_car.save('bama-car-{}.xlsx'.format(timenow))


        time.sleep(21600)



class myThread(threading.Thread):
	def __init__(self, threadID, name, counter, lock1, lock2, lock3):
		threading.Thread.__init__(self)
		self.threadID = threadID
		self.name = name
		self.counter = counter
		self.lock1 = lock1
		self.lock2 = lock2
		self.lock3 = lock3
	def run(self):
		runner()
	

if __name__ == "__main__":
    attack_threads = []

    for i in range(2):
        attack_threads.append(myThread(i, "Thread-{}".format(i), i, lock1, lock2, lock3))
        attack_threads[i].start()

    for i in range(2):
        attack_threads[i].join()


